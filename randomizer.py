"""A python script to rename files.  Needs Python 2 or 3, easygui, numpy,
and openpyxl (all pip installable as of Aug 26th 2019).
Beth Cimini, 2014-2019
"""

from openpyxl import Workbook, load_workbook
import os
import easygui as eg
from numpy import random

def file_name_randomizer():
    b_or_not_to_b = eg.choicebox(msg='Do you want to blind or unblind files?', choices=['Blind', 'Unblind'])
    if b_or_not_to_b == 'Blind':
        folder=eg.diropenbox('Where are the files you want to randomize?')
        files=os.listdir(folder)
        foldname=os.path.split(folder)[1]
        workbook=Workbook()
        name=foldname+" randomization"
        while len(name)>=31:
            name=eg.enterbox("Filename "+name+" is too long- enter a shorter one")
        randsheet = workbook.active
        randsheet.title = name
        rowcount=2
        numberlist=[]
        randsheet['A1'] = 'NewName'
        randsheet['B1'] = 'OriginalName'
        for i in files:
            number="%0*d" %(6,random.randint(0,len(files)*3))
            while number in numberlist:
                number="%0*d" %(6,random.randint(0,len(files)*3))
            randsheet['A'+str(rowcount)] = number+i[i.rindex('.'):]
            randsheet['B'+str(rowcount)] = i
            os.rename(os.path.join(folder,i),os.path.join(folder,number+i[i.rindex('.'):]))
            rowcount+=1
            numberlist.append(number)
        workbook.save(filename = os.path.join(folder,name+'.xlsx'))
    elif b_or_not_to_b == 'Unblind':
        folder,blinder = os.path.split(eg.fileopenbox(msg='Which is the blinding code file?', filetypes=['*.xlsx']))
        wb = load_workbook(filename=os.path.join(folder,blinder))
        randsheet = wb.active
        for row in range(2,randsheet.max_row+1):
            os.rename(os.path.join(folder,randsheet['A'+str(row)].value),os.path.join(folder,randsheet['B'+str(row)].value))
    else:
        pass

file_name_randomizer()
